"""
CSV/Excel import routes for MARIAM menus.

3-step workflow:
1. Upload  - parse the file, returns columns and preview
2. Preview - apply column mapping, returns menus to be created
3. Confirm - execute the import into the database

Endpoints (editor role required):
- POST /v1/imports/menus/upload
- POST /v1/imports/menus/preview
- POST /v1/imports/menus/confirm
"""
import csv
import io
import uuid
import re
from datetime import datetime, date, timedelta
from functools import wraps
from flask import request, jsonify, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from flask_smorest import Blueprint
from ..extensions import db
from ..models import User, Restaurant, Menu, MenuItem, AuditLog, ImportSession
from ..models import DietaryTag, Certification, DietaryTagKeyword, CertificationKeyword
from ..models.category import MenuCategory
from ..security import get_client_ip
from ..schemas.imports import ImportUploadSchema, ImportPreviewSchema, ImportConfirmSchema
from ..schemas.common import ErrorSchema


imports_bp = Blueprint(
    'imports', __name__,
    description='CSV/Excel import — Bulk menu loading from file'
)


# ============================================================
# HELPERS
# ============================================================

def editor_required(f):
    """Décorateur : accès réservé aux éditeurs (role editor ou admin)."""
    @wraps(f)
    @jwt_required()
    def decorated_function(*args, **kwargs):
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        if not user or not user.is_editor():
            return jsonify({'error': 'Accès réservé aux éditeurs'}), 403
        return f(*args, **kwargs)
    return decorated_function


def get_default_restaurant():
    return Restaurant.query.filter_by(is_active=True).first()


def detect_encoding(file_content: bytes) -> str:
    try:
        import chardet
        result = chardet.detect(file_content)
        return result.get('encoding', 'utf-8') or 'utf-8'
    except ImportError:
        for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                file_content.decode(encoding)
                return encoding
            except (UnicodeDecodeError, LookupError):
                continue
        return 'utf-8'


def detect_delimiter(sample: str) -> str:
    delimiters = [';', ',', '\t', '|']
    max_count = 0
    best_delimiter = ';'
    first_line = sample.split('\n')[0] if '\n' in sample else sample
    for delimiter in delimiters:
        count = first_line.count(delimiter)
        if count > max_count:
            max_count = count
            best_delimiter = delimiter
    return best_delimiter


def parse_csv_content(content: str, delimiter: str) -> tuple[list[str], list[dict]]:
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    columns = reader.fieldnames or []
    rows = list(reader)
    return columns, rows


def parse_excel_content(file_content: bytes) -> tuple[list[str], list[dict]]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return [], []
        columns = [str(h).strip() if h else f'Column_{i}' for i, h in enumerate(headers)]
        rows = []
        for row in rows_iter:
            row_dict = {columns[i]: str(v).strip() if v is not None else ''
                        for i, v in enumerate(row) if i < len(columns)}
            if any(row_dict.values()):
                rows.append(row_dict)
        wb.close()
        return columns, rows
    except ImportError:
        raise ValueError("La bibliothèque openpyxl n'est pas installée pour lire les fichiers Excel.")


def detect_date_format(date_str: str) -> str | None:
    formats = [
        ('%Y-%m-%d', 'YYYY-MM-DD'), ('%d/%m/%Y', 'DD/MM/YYYY'),
        ('%d-%m-%Y', 'DD-MM-YYYY'), ('%d.%m.%Y', 'DD.MM.YYYY'),
        ('%m/%d/%Y', 'MM/DD/YYYY'), ('%Y/%m/%d', 'YYYY/MM/DD'),
    ]
    for fmt, _ in formats:
        try:
            datetime.strptime(date_str.strip(), fmt)
            return fmt
        except ValueError:
            continue
    return None


def parse_date(date_str: str, date_format: str | None = None) -> date | None:
    if not date_str or not date_str.strip():
        return None
    date_str = date_str.strip()
    if date_format:
        try:
            return datetime.strptime(date_str, date_format).date()
        except ValueError:
            pass
    fmt = detect_date_format(date_str)
    if fmt:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            pass
    return None


def detect_tags_from_text(text: str) -> dict:
    text_lower = text.lower()
    detected_tag_ids: set[str] = set()
    detected_cert_ids: set[str] = set()
    for tk in DietaryTagKeyword.query.all():
        if tk.keyword in text_lower:
            detected_tag_ids.add(tk.tag_id)
    for ck in CertificationKeyword.query.all():
        if ck.keyword in text_lower:
            detected_cert_ids.add(ck.certification_id)
    return {'tags': sorted(detected_tag_ids), 'certifications': sorted(detected_cert_ids)}


def clean_item_name(name: str) -> str:
    for emoji in ['🌱', '🥬', '🥗', '🕌', '🌿', '🇫🇷', '♻️', '🐟', '🐄', '🐔']:
        name = name.replace(emoji, '')
    name = re.sub(r'\s*[\(\[](vg|végétarien|bio|halal|sans porc|local)[\)\]]', '', name, flags=re.IGNORECASE)
    return name.strip()


def normalize_label(label: str) -> str:
    """Normalise un label pour la comparaison (minuscules, sans accents)."""
    import unicodedata
    label = label.lower().strip()
    return ''.join(
        c for c in unicodedata.normalize('NFD', label)
        if unicodedata.category(c) != 'Mn'
    )


def suggest_column_mapping(columns: list[str], restaurant_id: int | None = None) -> dict:
    """Auto-mappe les colonnes CSV aux catégories DB par label (insensible à la casse/accents)."""
    mapping = {}
    date_patterns = ['date', 'jour', 'day', 'fecha']

    # Charger toutes les catégories du restaurant (principales + sous-catégories)
    categories: list[MenuCategory] = []
    if restaurant_id:
        categories = MenuCategory.query.filter_by(restaurant_id=restaurant_id).all()

    # Index label normalisé → category id (int)
    cat_label_index: dict[str, int] = {
        normalize_label(c.label): c.id for c in categories
    }

    for col in columns:
        col_norm = normalize_label(col)
        for pattern in date_patterns:
            if pattern in col_norm:
                mapping['date'] = col
                break
        for label_norm, cat_id in cat_label_index.items():
            if label_norm in col_norm or col_norm in label_norm:
                mapping.setdefault('categories', {})[col] = cat_id
                break
    return mapping


def build_menus_from_rows(rows, column_mapping, date_config, restaurant_id):
    menus = []
    date_column = None
    category_columns = {}

    for mapping in column_mapping:
        csv_col = mapping.get('csv_column')
        target = mapping.get('target_field')
        if target == 'date':
            date_column = csv_col
        elif target == 'category' and mapping.get('category_id'):
            category_columns[csv_col] = mapping['category_id']

    date_mode = date_config.get('mode', 'from_file')
    start_date_str = date_config.get('start_date')
    skip_weekends = date_config.get('skip_weekends', True)
    date_format = date_config.get('date_format')
    auto_detect_tags = date_config.get('auto_detect_tags', True)

    if date_mode in ['align_week', 'start_date'] and start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError(f"Format de date invalide: {start_date_str}")
    else:
        start_date = date.today()

    current_date = start_date

    for row in rows:
        if date_mode == 'from_file' and date_column:
            menu_date = parse_date(row.get(date_column, ''), date_format)
            if not menu_date:
                continue
        else:
            menu_date = current_date
            current_date = current_date + timedelta(days=1)
            while skip_weekends and current_date.weekday() >= 5:
                current_date = current_date + timedelta(days=1)

        items = []
        for csv_col, category_id in category_columns.items():
            cell_value = row.get(csv_col, '').strip()
            if not cell_value:
                continue
            for order, item_name in enumerate(re.split(r'[,\n]+', cell_value)):
                item_name = item_name.strip()
                if not item_name:
                    continue
                item = {
                    'category_id': category_id,  # integer FK
                    'name': clean_item_name(item_name),
                    'order': order,
                    'tags': [],
                    'certifications': [],
                }
                if auto_detect_tags:
                    detected = detect_tags_from_text(item_name)
                    item['tags'] = detected['tags']
                    item['certifications'] = detected['certifications']
                items.append(item)

        if items:
            menus.append({
                'date': menu_date.isoformat(),
                'date_display': menu_date.strftime('%A %d/%m/%Y'),
                'items': items,
                'has_duplicate': False,
                'existing_menu': None,
            })

    return menus


# ============================================================
# ENDPOINTS
# ============================================================

@imports_bp.route('/upload', methods=['POST'])
@imports_bp.response(200, ImportUploadSchema)
@imports_bp.alt_response(400, schema=ErrorSchema, description="Invalid or missing file")
@imports_bp.alt_response(500, schema=ErrorSchema, description="Processing error")
@editor_required
def upload_file():
    """Upload and parse a CSV or Excel file.

    Multipart/form-data request with `file` field.
    Accepted formats: `.csv`, `.xlsx`, `.xls`.

    Returns detected columns, a preview of the first 10 rows,
    an auto-suggested column mapping, and the detected date format.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Nom de fichier manquant'}), 400

    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        return jsonify({'error': 'Format non supporté. Utilisez CSV ou Excel (.xlsx)'}), 400

    try:
        file_content = file.read()

        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            columns, rows = parse_excel_content(file_content)
            delimiter = None
        else:
            encoding = detect_encoding(file_content)
            content = file_content.decode(encoding).replace('\r\n', '\n').replace('\r', '\n')
            delimiter = detect_delimiter(content)
            columns, rows = parse_csv_content(content, delimiter)

        if not columns:
            return jsonify({'error': 'Fichier vide ou impossible de lire les colonnes'}), 400
        if not rows:
            return jsonify({'error': 'Aucune donnée trouvée dans le fichier'}), 400

        file_id = str(uuid.uuid4())
        ImportSession.cleanup_expired()

        current_user_id = int(get_jwt_identity())
        session = ImportSession(
            id=file_id,
            user_id=current_user_id,
            filename=file.filename,
            columns=columns,
            rows=rows,
            expires_minutes=30,
        )
        db.session.add(session)
        db.session.commit()

        upload_restaurant_id = request.form.get('restaurant_id', type=int)
        if not upload_restaurant_id:
            default_restaurant = get_default_restaurant()
            if default_restaurant:
                upload_restaurant_id = default_restaurant.id
        auto_mapping = suggest_column_mapping(columns, upload_restaurant_id)

        detected_date_format = None
        if auto_mapping.get('date'):
            date_col = auto_mapping['date']
            for row in rows[:5]:
                date_val = row.get(date_col, '')
                if date_val:
                    detected_date_format = detect_date_format(date_val)
                    if detected_date_format:
                        break

        return jsonify({
            'file_id': file_id,
            'filename': file.filename,
            'columns': columns,
            'preview_rows': rows[:10],
            'row_count': len(rows),
            'detected_delimiter': delimiter,
            'auto_mapping': auto_mapping,
            'detected_date_format': detected_date_format,
        }), 200

    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        current_app.logger.error(f"Erreur upload CSV: {e}")
        return jsonify({'error': f'Erreur lors du traitement du fichier: {e}'}), 500


@imports_bp.route('/preview', methods=['POST'])
@imports_bp.arguments(ImportPreviewSchema)
@imports_bp.response(200, ImportUploadSchema)
@imports_bp.alt_response(400, schema=ErrorSchema, description="Invalid data")
@imports_bp.alt_response(404, schema=ErrorSchema, description="Session expired")
@editor_required
def preview_import(data):
    """Generate an import preview based on the provided column mapping.

    Returns the menus that will be created, flagging duplicates
    (existing menus for the same dates).
    """
    file_id = data.get('file_id')
    column_mapping = data.get('column_mapping', [])
    date_config = data.get('date_config', {})
    restaurant_id = data.get('restaurant_id')

    current_user_id = int(get_jwt_identity())
    session = ImportSession.get_valid(file_id, current_user_id)
    if not session:
        return jsonify({'error': 'Session expirée ou fichier non trouvé. Veuillez re-uploader le fichier.'}), 404

    rows = session.get_rows()

    if not restaurant_id:
        restaurant = get_default_restaurant()
        if restaurant:
            restaurant_id = restaurant.id

    try:
        menus = build_menus_from_rows(rows, column_mapping, date_config, restaurant_id)

        for menu_data in menus:
            existing = Menu.query.filter_by(
                restaurant_id=restaurant_id,
                date=datetime.strptime(menu_data['date'], '%Y-%m-%d').date()
            ).first()
            menu_data['has_duplicate'] = existing is not None
            if existing:
                menu_data['existing_menu'] = existing.to_dict()

        duplicates_count = sum(1 for m in menus if m['has_duplicate'])

        return jsonify({
            'menus': menus,
            'total_count': len(menus),
            'duplicates_count': duplicates_count,
            'new_count': len(menus) - duplicates_count,
        }), 200

    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        current_app.logger.error(f"Erreur preview import: {e}")
        return jsonify({'error': f'Erreur lors de la prévisualisation: {e}'}), 500


@imports_bp.route('/confirm', methods=['POST'])
@imports_bp.arguments(ImportConfirmSchema)
@imports_bp.response(200, ImportUploadSchema)
@imports_bp.alt_response(400, schema=ErrorSchema, description="Invalid data or missing restaurant")
@imports_bp.alt_response(404, schema=ErrorSchema, description="Session expired")
@editor_required
def confirm_import(data):
    """Confirm and execute the menu import into the database.

    - `duplicate_action`: `skip` (default), `replace` or `merge`
    - `auto_publish`: immediately publish the imported menus
    """
    current_user_id = int(get_jwt_identity())

    file_id = data.get('file_id')
    column_mapping = data.get('column_mapping', [])
    date_config = data.get('date_config', {})
    duplicate_action = data.get('duplicate_action', 'skip')
    auto_publish = data.get('auto_publish', False)
    restaurant_id = data.get('restaurant_id')

    session = ImportSession.get_valid(file_id, current_user_id)
    if not session:
        return jsonify({'error': 'Session expirée ou fichier non trouvé. Veuillez re-uploader le fichier.'}), 404

    rows = session.get_rows()

    if not restaurant_id:
        restaurant = get_default_restaurant()
        if restaurant:
            restaurant_id = restaurant.id
    if not restaurant_id:
        return jsonify({'error': 'Aucun restaurant configuré'}), 400

    try:
        menus_data = build_menus_from_rows(rows, column_mapping, date_config, restaurant_id)

        imported_count = replaced_count = skipped_count = 0

        for menu_data in menus_data:
            menu_date = datetime.strptime(menu_data['date'], '%Y-%m-%d').date()
            existing_menu = Menu.query.filter_by(restaurant_id=restaurant_id, date=menu_date).first()

            if existing_menu:
                if duplicate_action == 'skip':
                    skipped_count += 1
                    continue
                elif duplicate_action == 'replace':
                    MenuItem.query.filter_by(menu_id=existing_menu.id).delete()
                    menu = existing_menu
                    replaced_count += 1
                elif duplicate_action == 'merge':
                    menu = existing_menu
                    replaced_count += 1
                else:
                    continue
            else:
                menu = Menu(restaurant_id=restaurant_id, date=menu_date, status='draft')
                db.session.add(menu)
                db.session.flush()
                imported_count += 1

            for idx, item_data in enumerate(menu_data['items']):
                tag_ids = item_data.get('tags', [])
                cert_ids = item_data.get('certifications', [])
                item = MenuItem(
                    menu_id=menu.id,
                    category_id=item_data['category_id'],
                    name=item_data['name'],
                    order=item_data.get('order', idx),
                )
                if tag_ids:
                    item.tags = DietaryTag.query.filter(DietaryTag.id.in_(tag_ids)).all()
                if cert_ids:
                    item.certifications = Certification.query.filter(Certification.id.in_(cert_ids)).all()
                db.session.add(item)

            if auto_publish:
                menu.status = 'published'
                menu.published_at = datetime.utcnow()
                menu.published_by_id = current_user_id

        AuditLog.log(
            action='csv_import',
            user_id=current_user_id,
            details={
                'filename': session.filename,
                'imported_count': imported_count,
                'replaced_count': replaced_count,
                'skipped_count': skipped_count,
                'auto_publish': auto_publish,
            },
            ip_address=get_client_ip()
        )

        db.session.commit()

        db.session.delete(session)
        db.session.commit()

        return jsonify({
            'success': True,
            'imported_count': imported_count,
            'replaced_count': replaced_count,
            'skipped_count': skipped_count,
            'message': f'{imported_count + replaced_count} menu(s) importé(s)',
        }), 200

    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur import CSV: {e}")
        return jsonify({'error': f"Erreur lors de l'import: {e}"}), 500
