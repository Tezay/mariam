"""
Routes d'import CSV/Excel pour les menus MARIAM.

Endpoints :
- POST /api/menus/import/upload - Upload et parsing initial
- POST /api/menus/import/preview - Prévisualisation avec mapping
- POST /api/menus/import/confirm - Confirmation et import final
"""
import csv
import io
import uuid
import re
from datetime import datetime, date, timedelta
from functools import wraps
from flask import Blueprint, request, jsonify, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from ..extensions import db
from ..models import User, Restaurant, Menu, MenuItem, AuditLog, ImportSession
from ..security import get_client_ip

csv_import_bp = Blueprint('csv_import', __name__)


def editor_required(f):
    """Décorateur pour protéger les routes editor+."""
    @wraps(f)
    @jwt_required()
    def decorated_function(*args, **kwargs):
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user or not user.is_editor():
            return jsonify({'error': 'Accès réservé aux éditeurs'}), 403
        
        return f(*args, **kwargs)
    return decorated_function


def detect_encoding(file_content: bytes) -> str:
    """Détecte l'encodage du fichier."""
    try:
        import chardet
        result = chardet.detect(file_content)
        return result.get('encoding', 'utf-8') or 'utf-8'
    except ImportError:
        # Fallback si chardet n'est pas installé
        for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                file_content.decode(encoding)
                return encoding
            except (UnicodeDecodeError, LookupError):
                continue
        return 'utf-8'


def detect_delimiter(sample: str) -> str:
    """Détecte le délimiteur CSV."""
    delimiters = [';', ',', '\t', '|']
    max_count = 0
    best_delimiter = ';'
    
    first_line = sample.split('\n')[0] if '\n' in sample else sample
    
    for delimiter in delimiters:
        count = first_line.count(delimiter)
        if count > max_count:
            max_count = count
            best_delimiter = delimiter
    
    return best_delimiter


def parse_csv_content(content: str, delimiter: str) -> tuple[list[str], list[dict]]:
    """Parse le contenu CSV et retourne les colonnes et les lignes."""
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    columns = reader.fieldnames or []
    rows = list(reader)
    return columns, rows


def parse_excel_content(file_content: bytes) -> tuple[list[str], list[dict]]:
    """Parse un fichier Excel (.xlsx) et retourne les colonnes et les lignes."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        
        rows_iter = ws.iter_rows(values_only=True)
        
        # Première ligne = headers
        headers = next(rows_iter, None)
        if not headers:
            return [], []
        
        columns = [str(h).strip() if h else f'Column_{i}' for i, h in enumerate(headers)]
        
        # Lignes de données
        rows = []
        for row in rows_iter:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(columns):
                    row_dict[columns[i]] = str(value).strip() if value is not None else ''
            if any(row_dict.values()):  # Ignorer les lignes vides
                rows.append(row_dict)
        
        wb.close()
        return columns, rows
        
    except ImportError:
        raise ValueError("La bibliothèque openpyxl n'est pas installée pour lire les fichiers Excel.")


def detect_date_format(date_str: str) -> str | None:
    """Tente de détecter le format d'une date."""
    formats = [
        ('%Y-%m-%d', 'YYYY-MM-DD'),
        ('%d/%m/%Y', 'DD/MM/YYYY'),
        ('%d-%m-%Y', 'DD-MM-YYYY'),
        ('%d.%m.%Y', 'DD.MM.YYYY'),
        ('%m/%d/%Y', 'MM/DD/YYYY'),
        ('%Y/%m/%d', 'YYYY/MM/DD'),
    ]
    
    for fmt, label in formats:
        try:
            datetime.strptime(date_str.strip(), fmt)
            return fmt
        except ValueError:
            continue
    
    return None


def parse_date(date_str: str, date_format: str | None = None) -> date | None:
    """Parse une date avec le format donné ou détection automatique."""
    if not date_str or not date_str.strip():
        return None
    
    date_str = date_str.strip()
    
    if date_format:
        try:
            return datetime.strptime(date_str, date_format).date()
        except ValueError:
            pass
    
    # Détection automatique
    fmt = detect_date_format(date_str)
    if fmt:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            pass
    
    return None


def detect_tags_from_text(text: str) -> dict:
    """Détecte les tags alimentaires à partir du texte."""
    text_lower = text.lower()
    tags = []
    certifications = []
    
    # Tags alimentaires
    vegetarian_keywords = [
        'végétarien', 'vegetarien', 'vegan', 'végan', 'vg', 'veggie', 'sans viande',
        '🌱', '🥬', '🥗', '🥦'
    ]
    halal_keywords = ['halal', '🕌', 'hl']
    pork_free_keywords = ['sans porc', 'sans-porc', 'no pork', 'sp', 'volaille', 'dinde', 'poulet']
    gluten_free_keywords = ['sans gluten', 'gluten-free', 'gluten free', 'sg', 'gf']
    lactose_free_keywords = ['sans lactose', 'lactose-free', 'lactose free', 'sl', 'lf']
    
    for kw in vegetarian_keywords:
        if kw in text_lower:
            tags.append('vegetarian')
            break
    
    for kw in halal_keywords:
        if kw in text_lower:
            tags.append('halal')
            break
    
    for kw in pork_free_keywords:
        if kw in text_lower:
            tags.append('pork_free')
            break
    
    for kw in gluten_free_keywords:
        if kw in text_lower:
            tags.append('gluten_free')
            break
    
    for kw in lactose_free_keywords:
        if kw in text_lower:
            tags.append('lactose_free')
            break
    
    # Certifications
    bio_keywords = ['bio', 'biologique', 'organic', 'ab', 'agriculture biologique', '🌿']
    local_keywords = ['local', 'locaux', 'région', 'régional', 'circuit court', 'fermier']
    french_keywords = ['france', 'français', 'française', 'viande française', 'vf', 'volaille française', '🇫🇷', 'origine france']
    
    for kw in bio_keywords:
        if kw in text_lower:
            certifications.append('bio')
            break
    
    for kw in local_keywords:
        if kw in text_lower:
            certifications.append('local')
            break
    
    for kw in french_keywords:
        if kw in text_lower:
            certifications.append('french_meat')
            break
    
    return {
        'tags': list(set(tags)),
        'certifications': list(set(certifications))
    }


def clean_item_name(name: str) -> str:
    """Nettoie le nom d'un item en retirant les tags inline."""
    # Retirer les emojis courants
    emojis = ['🌱', '🥬', '🥗', '🕌', '🌿', '🇫🇷', '♻️', '🐟', '🐄', '🐔']
    for emoji in emojis:
        name = name.replace(emoji, '')
    
    # Retirer les mentions de tags entre parenthèses ou crochets
    name = re.sub(r'\s*[\(\[](vg|végétarien|bio|halal|sans porc|local)[\)\]]', '', name, flags=re.IGNORECASE)
    
    return name.strip()


def get_default_restaurant():
    """Retourne le restaurant par défaut."""
    return Restaurant.query.filter_by(is_active=True).first()


@csv_import_bp.route('/upload', methods=['POST'])
@editor_required
def upload_file():
    """
    Upload et parse un fichier CSV ou Excel.
    
    Retourne les colonnes détectées et un aperçu des données.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    
    file = request.files['file']
    
    if not file.filename:
        return jsonify({'error': 'Nom de fichier manquant'}), 400
    
    filename = file.filename.lower()
    
    # Vérifier le format
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        return jsonify({'error': 'Format non supporté. Utilisez CSV ou Excel (.xlsx)'}), 400
    
    try:
        file_content = file.read()
        
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            columns, rows = parse_excel_content(file_content)
            delimiter = None
        else:
            # CSV
            encoding = detect_encoding(file_content)
            content = file_content.decode(encoding)
            
            # Normaliser les fins de ligne
            content = content.replace('\r\n', '\n').replace('\r', '\n')
            
            delimiter = detect_delimiter(content)
            columns, rows = parse_csv_content(content, delimiter)
        
        if not columns:
            return jsonify({'error': 'Fichier vide ou impossible de lire les colonnes'}), 400
        
        if not rows:
            return jsonify({'error': 'Aucune donnée trouvée dans le fichier'}), 400
        
        # Générer un ID unique pour ce fichier
        file_id = str(uuid.uuid4())
        
        # Nettoyer les sessions expirées
        ImportSession.cleanup_expired()
        
        # Stocker les données parsées en base de données
        current_user_id = int(get_jwt_identity())
        session = ImportSession(
            id=file_id,
            user_id=current_user_id,
            filename=file.filename,
            columns=columns,
            rows=rows,
            expires_minutes=30
        )
        db.session.add(session)
        db.session.commit()
        
        # Essayer de détecter automatiquement les correspondances
        auto_mapping = suggest_column_mapping(columns)
        
        # Détecter le format de date si une colonne date est trouvée
        detected_date_format = None
        if auto_mapping.get('date'):
            date_col = auto_mapping['date']
            for row in rows[:5]:  # Vérifier les 5 premières lignes
                date_val = row.get(date_col, '')
                if date_val:
                    detected_date_format = detect_date_format(date_val)
                    if detected_date_format:
                        break
        
        return jsonify({
            'file_id': file_id,
            'filename': file.filename,
            'columns': columns,
            'preview_rows': rows[:10],  # Aperçu des 10 premières lignes
            'row_count': len(rows),
            'detected_delimiter': delimiter,
            'auto_mapping': auto_mapping,
            'detected_date_format': detected_date_format
        }), 200
        
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        current_app.logger.error(f"Erreur upload CSV: {str(e)}")
        return jsonify({'error': f'Erreur lors du traitement du fichier: {str(e)}'}), 500


def suggest_column_mapping(columns: list[str]) -> dict:
    """Suggère un mapping automatique basé sur les noms de colonnes."""
    mapping = {}
    
    # Patterns pour la date
    date_patterns = ['date', 'jour', 'day', 'fecha']
    
    # Patterns pour les catégories
    category_patterns = {
        'entree': ['entrée', 'entree', 'starter', 'appetizer', 'hors d\'oeuvre'],
        'plat': ['plat', 'plat principal', 'main', 'main course', 'dish', 'principal'],
        'vg': ['vg', 'végétarien', 'vegetarien', 'végé', 'vege', 'vegan', 'vegetarian'],
        'dessert': ['dessert', 'sweet', 'postre']
    }
    
    for col in columns:
        col_lower = col.lower().strip()
        
        # Chercher une colonne date
        for pattern in date_patterns:
            if pattern in col_lower:
                mapping['date'] = col
                break
        
        # Chercher les catégories
        for category_id, patterns in category_patterns.items():
            for pattern in patterns:
                if pattern in col_lower:
                    if 'categories' not in mapping:
                        mapping['categories'] = {}
                    mapping['categories'][col] = category_id
                    break
    
    return mapping



@csv_import_bp.route('/preview', methods=['POST'])
@editor_required
def preview_import():
    """
    Génère un aperçu de l'import basé sur le mapping fourni.
    """
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'Données manquantes'}), 400
    
    file_id = data.get('file_id')
    column_mapping = data.get('column_mapping', [])
    date_config = data.get('date_config', {})
    restaurant_id = data.get('restaurant_id')
    
    # Récupérer la session d'import depuis la base de données
    current_user_id = int(get_jwt_identity())
    session = ImportSession.get_valid(file_id, current_user_id)
    
    if not session:
        return jsonify({'error': 'Session expirée ou fichier non trouvé. Veuillez re-uploader le fichier.'}), 404
    
    rows = session.get_rows()
    
    # Récupérer le restaurant
    if not restaurant_id:
        restaurant = get_default_restaurant()
        if restaurant:
            restaurant_id = restaurant.id
    
    try:
        menus = build_menus_from_rows(rows, column_mapping, date_config, restaurant_id)
        
        # Vérifier les doublons
        for menu_data in menus:
            existing = Menu.query.filter_by(
                restaurant_id=restaurant_id,
                date=menu_data['date']
            ).first()
            
            menu_data['has_duplicate'] = existing is not None
            if existing:
                menu_data['existing_menu'] = existing.to_dict()
        
        # Compter les statistiques
        duplicates_count = sum(1 for m in menus if m['has_duplicate'])
        
        return jsonify({
            'menus': menus,
            'total_count': len(menus),
            'duplicates_count': duplicates_count,
            'new_count': len(menus) - duplicates_count
        }), 200
        
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        current_app.logger.error(f"Erreur preview import: {str(e)}")
        return jsonify({'error': f'Erreur lors de la prévisualisation: {str(e)}'}), 500


def build_menus_from_rows(rows: list[dict], column_mapping: list[dict], 
                          date_config: dict, restaurant_id: int) -> list[dict]:
    """Construit les menus à partir des lignes CSV et du mapping."""
    menus = []
    
    # Extraire le mapping
    date_column = None
    category_columns = {}  # {csv_column: category_id}
    
    for mapping in column_mapping:
        csv_col = mapping.get('csv_column')
        target = mapping.get('target_field')
        
        if target == 'date':
            date_column = csv_col
        elif target == 'category' and mapping.get('category_id'):
            category_columns[csv_col] = mapping['category_id']
    
    # Configuration des dates
    date_mode = date_config.get('mode', 'from_file')
    start_date_str = date_config.get('start_date')
    skip_weekends = date_config.get('skip_weekends', True)
    date_format = date_config.get('date_format')
    
    # Date de départ pour les modes align/start_date
    if date_mode in ['align_week', 'start_date'] and start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError(f"Format de date invalide: {start_date_str}")
    else:
        start_date = date.today()
    
    current_date = start_date
    auto_detect_tags = date_config.get('auto_detect_tags', True)
    
    for row_idx, row in enumerate(rows):
        # Déterminer la date
        if date_mode == 'from_file' and date_column:
            date_str = row.get(date_column, '')
            menu_date = parse_date(date_str, date_format)
            if not menu_date:
                continue  # Ignorer les lignes sans date valide
        else:
            # Mode séquentiel
            menu_date = current_date
            
            # Avancer au prochain jour (en sautant les week-ends si demandé)
            current_date = current_date + timedelta(days=1)
            while skip_weekends and current_date.weekday() >= 5:
                current_date = current_date + timedelta(days=1)
        
        # Créer les items
        items = []
        for csv_col, category_id in category_columns.items():
            cell_value = row.get(csv_col, '').strip()
            
            if not cell_value:
                continue
            
            # Possibilité de plusieurs items séparés par des virgules ou retours à la ligne
            item_names = re.split(r'[,\n]+', cell_value)
            
            for order, item_name in enumerate(item_names):
                item_name = item_name.strip()
                if not item_name:
                    continue
                
                item = {
                    'category': category_id,
                    'name': clean_item_name(item_name),
                    'order': order,
                    'tags': [],
                    'certifications': []
                }
                
                # Marquer automatiquement VG comme végétarien
                if category_id == 'vg':
                    item['is_vegetarian'] = True
                    item['tags'].append('vegetarian')
                
                # Détection automatique des tags
                if auto_detect_tags:
                    detected = detect_tags_from_text(item_name)
                    item['tags'] = list(set(item['tags'] + detected['tags']))
                    item['certifications'] = detected['certifications']
                
                items.append(item)
        
        if items:  # Seulement ajouter si on a des items
            menus.append({
                'date': menu_date.isoformat(),
                'date_display': menu_date.strftime('%A %d/%m/%Y'),
                'items': items,
                'has_duplicate': False,
                'existing_menu': None
            })
    
    return menus


@csv_import_bp.route('/confirm', methods=['POST'])
@editor_required
def confirm_import():
    """
    Confirme et exécute l'import des menus.
    """
    current_user_id = int(get_jwt_identity())
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'Données manquantes'}), 400
    
    file_id = data.get('file_id')
    column_mapping = data.get('column_mapping', [])
    date_config = data.get('date_config', {})
    duplicate_action = data.get('duplicate_action', 'skip')  # skip, replace, merge
    auto_publish = data.get('auto_publish', False)
    restaurant_id = data.get('restaurant_id')
    
    # Récupérer la session d'import depuis la base de données
    current_user_id = int(get_jwt_identity())
    session = ImportSession.get_valid(file_id, current_user_id)
    
    if not session:
        return jsonify({'error': 'Session expirée ou fichier non trouvé. Veuillez re-uploader le fichier.'}), 404
    
    rows = session.get_rows()
    
    # Récupérer le restaurant
    if not restaurant_id:
        restaurant = get_default_restaurant()
        if restaurant:
            restaurant_id = restaurant.id
    
    if not restaurant_id:
        return jsonify({'error': 'Aucun restaurant configuré'}), 400
    
    try:
        menus_data = build_menus_from_rows(rows, column_mapping, date_config, restaurant_id)
        
        imported_count = 0
        replaced_count = 0
        skipped_count = 0
        
        for menu_data in menus_data:
            menu_date = datetime.strptime(menu_data['date'], '%Y-%m-%d').date()
            
            # Chercher un menu existant
            existing_menu = Menu.query.filter_by(
                restaurant_id=restaurant_id,
                date=menu_date
            ).first()
            
            if existing_menu:
                if duplicate_action == 'skip':
                    skipped_count += 1
                    continue
                elif duplicate_action == 'replace':
                    # Supprimer les anciens items
                    MenuItem.query.filter_by(menu_id=existing_menu.id).delete()
                    menu = existing_menu
                    replaced_count += 1
                elif duplicate_action == 'merge':
                    # Garder le menu existant et ajouter les items
                    menu = existing_menu
                    replaced_count += 1
                else:
                    continue
            else:
                # Créer un nouveau menu
                menu = Menu(
                    restaurant_id=restaurant_id,
                    date=menu_date,
                    status='draft'
                )
                db.session.add(menu)
                db.session.flush()
                imported_count += 1
            
            # Ajouter les items
            for idx, item_data in enumerate(menu_data['items']):
                item = MenuItem(
                    menu_id=menu.id,
                    category=item_data['category'],
                    name=item_data['name'],
                    order=item_data.get('order', idx),
                    is_vegetarian='vegetarian' in item_data.get('tags', []),
                    is_halal='halal' in item_data.get('tags', []),
                    is_pork_free='pork_free' in item_data.get('tags', []),
                    tags=item_data.get('tags', []),
                    certifications=item_data.get('certifications', [])
                )
                db.session.add(item)
            
            # Publier si demandé
            if auto_publish:
                menu.status = 'published'
                menu.published_at = datetime.utcnow()
                menu.published_by_id = current_user_id
        
        # Logger l'action
        AuditLog.log(
            action='csv_import',
            user_id=current_user_id,
            details={
                'filename': session.filename,
                'imported_count': imported_count,
                'replaced_count': replaced_count,
                'skipped_count': skipped_count,
                'auto_publish': auto_publish
            },
            ip_address=get_client_ip()
        )
        
        db.session.commit()
        
        # Supprimer la session d'import
        db.session.delete(session)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'imported_count': imported_count,
            'replaced_count': replaced_count,
            'skipped_count': skipped_count,
            'message': f'{imported_count + replaced_count} menu(s) importé(s)'
        }), 200
        
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur import CSV: {str(e)}")
        return jsonify({'error': f'Erreur lors de l\'import: {str(e)}'}), 500
