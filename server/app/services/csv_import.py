"""
Helpers partagés d'import CSV/Excel — parsing de fichier et détection de taxonomie.

Utilisés par l'import de menus et l'import du catalogue de plats (routes/imports.py).
Le parsing est générique (encodage, délimiteur, CSV, Excel) ; la détection de tags
s'appuie sur les mots-clés de taxonomie stockés en base.
"""
import csv
import io
import re
import unicodedata

from ..models import CertificationKeyword, DietaryTagKeyword


def detect_encoding(file_content: bytes) -> str:
    """Détecte l'encodage d'un fichier (chardet si disponible, sinon essais successifs)."""
    try:
        import chardet
        result = chardet.detect(file_content)
        return result.get('encoding', 'utf-8') or 'utf-8'
    except ImportError:
        for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                file_content.decode(encoding)
                return encoding
            except (UnicodeDecodeError, LookupError):
                continue
        return 'utf-8'


def detect_delimiter(sample: str) -> str:
    """Devine le délimiteur d'un CSV à partir de la première ligne."""
    delimiters = [';', ',', '\t', '|']
    max_count = 0
    best_delimiter = ';'
    first_line = sample.split('\n')[0] if '\n' in sample else sample
    for delimiter in delimiters:
        count = first_line.count(delimiter)
        if count > max_count:
            max_count = count
            best_delimiter = delimiter
    return best_delimiter


def parse_csv_content(content: str, delimiter: str) -> tuple[list[str], list[dict]]:
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    columns = reader.fieldnames or []
    rows = list(reader)
    return columns, rows


def parse_excel_content(file_content: bytes) -> tuple[list[str], list[dict]]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return [], []
        columns = [str(h).strip() if h else f'Column_{i}' for i, h in enumerate(headers)]
        rows = []
        for row in rows_iter:
            row_dict = {columns[i]: str(v).strip() if v is not None else ''
                        for i, v in enumerate(row) if i < len(columns)}
            if any(row_dict.values()):
                rows.append(row_dict)
        wb.close()
        return columns, rows
    except ImportError:
        raise ValueError("La bibliothèque openpyxl n'est pas installée pour lire les fichiers Excel.") from None


def parse_upload(file) -> tuple[list[str], list[dict], str | None]:
    """Parse un fichier uploadé (CSV ou Excel) et retourne (colonnes, lignes, délimiteur).

    Le délimiteur vaut None pour les fichiers Excel. Lève ValueError si le format
    n'est pas supporté ou si le fichier est vide.
    """
    filename = (file.filename or '').lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise ValueError('Format non supporté. Utilisez CSV ou Excel (.xlsx)')

    file_content = file.read()
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        columns, rows = parse_excel_content(file_content)
        delimiter = None
    else:
        encoding = detect_encoding(file_content)
        content = file_content.decode(encoding).replace('\r\n', '\n').replace('\r', '\n')
        delimiter = detect_delimiter(content)
        columns, rows = parse_csv_content(content, delimiter)

    if not columns:
        raise ValueError('Fichier vide ou impossible de lire les colonnes')
    if not rows:
        raise ValueError('Aucune donnée trouvée dans le fichier')

    return columns, rows, delimiter


def normalize_label(label: str) -> str:
    """Normalise un label pour la comparaison (minuscules, sans accents)."""
    label = label.lower().strip()
    return ''.join(
        c for c in unicodedata.normalize('NFD', label)
        if unicodedata.category(c) != 'Mn'
    )


def detect_tags_from_text(text: str) -> dict:
    """Détecte tags alimentaires et certifications dans un texte via les mots-clés DB."""
    text_lower = text.lower()
    detected_tag_ids: set[str] = set()
    detected_cert_ids: set[str] = set()
    for tk in DietaryTagKeyword.query.all():
        if tk.keyword in text_lower:
            detected_tag_ids.add(tk.tag_id)
    for ck in CertificationKeyword.query.all():
        if ck.keyword in text_lower:
            detected_cert_ids.add(ck.certification_id)
    return {'tags': sorted(detected_tag_ids), 'certifications': sorted(detected_cert_ids)}


def clean_item_name(name: str) -> str:
    """Nettoie un nom de plat issu d'un CSV (retire emojis et suffixes de régime)."""
    for emoji in ['🌱', '🥬', '🥗', '🕌', '🌿', '🇫🇷', '♻️', '🐟', '🐄', '🐔']:
        name = name.replace(emoji, '')
    name = re.sub(r'\s*[\(\[](vg|végétarien|bio|halal|sans porc|local)[\)\]]', '', name, flags=re.IGNORECASE)
    return name.strip()
